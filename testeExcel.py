#sheets - são as paginas de uma planilha
#workbook - é um arquivo que contem estas paginas
import openpyxl
planilha_teste = openpyxl.Workbook()
planilha_teste.create_sheet("frutas")
planilha_teste.create_sheet("legumes")
planilha_teste.create_sheet("sementes")
print(planilha_teste.sheetnames)

cabecalho_frutas = ["Titulo","Localização", "Preço"]
sheet_frutas = planilha_teste.get_sheet_by_name("frutas")
sheet_frutas.append(cabecalho_frutas)

frutas = [['Uva','mercado',15],['Maça','Padaria',15],['cafe','ultramercado',20]]
for fruta in frutas:
    sheet_frutas.append(fruta)


planilha_teste.save("Dados Supermercado.xlsx")