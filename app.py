import openpyxl
planilha = openpyxl.load_workbook("pessoas.xlsx")
#printar o nome dos sheeets
print(planilha.sheetnames)

#pegar o valor e alterar na planilha
sheet1 = planilha.get_sheet_by_name('Sheet1')
print(sheet1['C3'].value)
sheet1['C3'].value = "Leal" #alterar valor  
print(sheet1['C3'].value)
#outra maneira
sheet1.cell(row=3,column=3,value="kakashi")#alterar valor
print( sheet1['C3'].value) #imprimir valor na tela

#percorrer planilha / linha
for linha in sheet1.iter_rows(min_row=2,max_row=10,min_col=2,max_col=4):
    print(linha[0].value,linha[1].value,linha[2].value)

#percorre planilha / coluna
for linha in sheet1.iter_cols(min_col=2,max_col=2,min_row=2):
    for cell in linha:
        print(cell.value)