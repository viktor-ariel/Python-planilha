import openpyxl

planilha = openpyxl.load_workbook("pessoas.xlsx")
print(planilha.sheetnames)
sheet1 = planilha.get_sheet_by_name('Sheet1')

#Desafio 1
print(sheet1['B4'].value)

#Desfio 2 
#1 forma
sheet1['B4'] = "Jimmy"
#2 forma
sheet1.cell(row=4,column=2,value="Viktor")
print(sheet1['B4'].value)

#Desafio 3
for linha in sheet1.iter_rows(min_row=2, max_row=11):
    print(linha[0].value,linha[1].value,linha[2].value,linha[3].value,linha[4].value,
          linha[5].value,linha[6].value,linha[7].value)

#Desafio 4
for linha in sheet1.iter_cols(min_col=3, max_col=3, min_row=1):
    for celula in linha:
        print(celula.value)




